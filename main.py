# import openpyxl module 
import openpyxl 
import base64
import email
import imaplib
import os
import re
import shutil
import jpype
import asposecells
import win32com.client as win32
from openpyxl.styles import Alignment
from datetime import datetime
from datetime import timedelta
from math import *
from datetime import datetime
import re
import io
import pandas as pd
import asposecells
import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell
from openpyxl.utils import get_column_letter
import xlrd    
  
from pdfminer.converter import TextConverter
from pdfminer.pdfinterp import PDFPageInterpreter
from pdfminer.pdfinterp import PDFResourceManager
from pdfminer.pdfpage import PDFPage


class Info:
    def __init__(self):
        filename = 'Дисциплины.txt'
        global listt
        global group
        global hours 
        global data
        self.list = Info.get_list(self, filename)
        self.group = Info.get_group(self, filename)
        self.hours = Info.get_hours(self, filename)
        self.data = Info.get_date(self, filename)
         
        wb = openpyxl.Workbook()
        for j in range(len(self.hours)):
            if int(self.hours[j][0])!=0:
                Info.creat_list_lsr(self, wb, 'лек', j)
            if (int(self.hours[j][1]!=0)) and (int(self.hours[j][2]!=0)):
                Info.creat_list_sr(self,wb, j)
            elif int(self.hours[j][1])!=0:
                Info.creat_list_lsr(self, wb, 'сем', j)
            elif int(self.hours[j][2])!=0:
                Info.creat_list_lsr(self, wb, 'лр', j)
        del wb['Sheet']
        wb.save("sample.xlsx")
        listt = self.list
        group = self.group
        hours = self.hours 
        data = self.data

        # Info.extract_text(self, wb)
   
    def get_list(self, filename):
        subject_list =[]
        gr = ' СМ5-'
        lec = 'лек '
        cem = 'сем '
        lr = 'лр '
        with open(filename, 'r', encoding='UTF-8') as file:
            while line := file.readline():
                n = line.count(gr)
                if lec in line:
                    if line.index(gr)>line.index(lec):
                        ind = line.index(lec) 
                    else: ind = line.index(gr)  
                elif cem in line:
                    if line.index(gr)>line.index(cem):
                        ind = line.index(cem)
                    else: ind = line.index(gr) 
                else: 
                    if line.index(gr)>line.index(lr):
                        ind = line.index(lr)
                    else: ind = line.index(gr)
                delete = line[ind:len(line)]
                list = line.replace(delete, '')
                for i in range(n):
                    subject_list.append(list)
        for dir in subject_list:
            tmp = os.path.join(os.getcwd(),dir)
            if os.path.exists(tmp):
                pass    
            else:
                os.mkdir(tmp)
                print(f"Директория {tmp} создана")
        
        return subject_list

    def get_group(self, filename):
        subject_group =[]
        gr = ' СМ5-'
        with open(filename, 'r', encoding='UTF-8') as file:
            while line := file.readline():
                list = line
                while gr in list:
                    ind = list.index(gr)    
                    idend = ind+8
                    subject_group.append(list[ind+1:idend].replace(' ', ''))
                    delete = list[ind:idend]
                    list = list.replace(delete, '')
        return subject_group
    
    def get_hours(self, filename):
        hours = []
        subject_hours = []
        lec = 'лек '
        cem = 'сем '
        lr = 'лр '
        gr = ' СМ5-'
        with open(filename, 'r', encoding='UTF-8') as file:
            while line := file.readline():
                n = line.count(gr)
                list = line
                if lec in list:
                    ind = list.index(lec)       
                    hours.append(int(list[ind+3+1:ind+3+3]))
                    delete = lec+list[ind+3+1:ind+3+3]
                    list = list.replace(delete, '')
                else:
                    hours.append(int('0'))
                if cem in list:
                    ind = list.index(cem)       
                    hours.append(int(list[ind+3+1:ind+3+3]))
                    delete = cem+list[ind+3+1:ind+3+3]
                    list = list.replace(delete, '')
                else:
                    hours.append(int('0'))
                if lr in list:
                    ind = list.index(lr)       
                    hours.append(int(list[ind+2+1:ind+2+3]))
                    delete = lr+list[ind+2+1:ind+2+3]
                    list = list.replace(delete, '')
                else:
                    hours.append(int('0'))
                for i in range(n):
                    subject_hours.append(hours)
                hours = []
        return subject_hours

    def get_date(self, filename):
        subject_date = []
        data = []
        date = '.'
        gr = ' СМ5-'
        i = 0
        with open(filename, 'r', encoding='UTF-8') as file:
            while line := file.readline():
                n = line.count(gr)
                list = line
                while date in list:
                    for j in range(len(self.hours[i])):
                        if int(self.hours[i][j]) !=0:
                            ind = list.index(date)
                            str = list[ind-2:ind+8].split('.')
                            data.append(str[0] + '-' + str[1] + '-' + str[2]) 
                            delete = list[ind-2:ind+8]
                            list = list.replace(delete, '', 1) 
                    subject_date.append(data)
                    data = []
                    i = i+1
        
                
        return subject_date

    def beatifule(sheet):
        sheet.column_dimensions['A'].width = 4
        sheet.column_dimensions['B'].width = 20
        sheet['A2'].value = "№"
        sheet['A2'].alignment = Alignment(horizontal="center", vertical="center")
        sheet['B2'].value = "ФИО"
        sheet['B2'].alignment = Alignment(horizontal="center", vertical="center")

    def creat_list_lsr(self, wb, litter, j):
        sheet = wb.create_sheet(str(self.group[j]) + litter)
        Info.beatifule(sheet)
        if litter == 'лек': i = 0
        elif litter == 'сем': i =1
        else: i = 2
        if self.hours[j][i] == 51:  hours = 17
        elif self.hours[j][i]  == 17:   hours = 8
        elif self.hours[j][i]  == 34:   hours = 8
        dim = 'A1:'+chr(67+hours)+'1'
        sheet.merge_cells(dim) 
        sheet.cell(row = 1, column = 1).value = self.group[j]+" "+self.list[j] +litter
        sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
        M = chr(67+hours)+'2'
        sheet[M].value = "M%"
        sheet[M].alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[chr(67+hours)].width = int(5)
        Begindate = datetime.strptime(self.data[j][0], "%d-%m-%Y" )
        Enddate = Begindate
        for i in range(hours):
            M = chr(67+i)+'2'
            sheet[M].value = str(Enddate.strftime("%d/%m"))[0:10]
            sheet[M].alignment = Alignment(horizontal="center", vertical="center", textRotation=90)
            sheet.column_dimensions[chr(67+i)].width = int(5)
            # sheet.row_dimensions['2'].height  = int(42)
            Enddate = Begindate + timedelta(days=7)
            Begindate = Enddate
        filename = self.group[j] +'.txt'
        group_list, group_dir = Info.creat_groups(filename)
        if litter != 'лр':
            for num in range(len(group_list)): 
                sheet.cell(row= num+3 , column = 1).value = num+1
                sheet.cell(row= num+3 , column = 2).value = group_list[num]
                summ = '= 100*COUNTA(C'+str(num+3) +':' + chr(67+hours-1) + str(num+3)+')'+'/'+str(hours)
                sheet.cell(row= num+3 , column = 3+hours).value = summ
                sheet.cell(row= num+3 , column = 1).alignment = Alignment(horizontal="center", vertical="center")
                sheet.cell(row= num+3 , column = 2).alignment = Alignment(horizontal="center", vertical="center")
                sheet.cell(row= num+3 , column = 3+hours).alignment = Alignment(horizontal="center", vertical="center")
        else:
            for num in range(ceil(len(group_list)/2)): 
                sheet.cell(row= num+3 , column = 1).value = num+1
                sheet.cell(row= num+3 , column = 2).value = group_list[num]
                summ = '= 100*COUNTA(C'+str(num+3) +':' + chr(67+hours-1) + str(num+3)+')'+'/'+str(hours)
                # summ = '= SUM(C3:' + chr(67+hours-1) + '3)'
                sheet.cell(row= num+3 , column = 3+hours).value = summ
                sheet.cell(row= num+3 , column = 1).alignment = Alignment(horizontal="center", vertical="center")
                sheet.cell(row= num+3 , column = 2).alignment = Alignment(horizontal="center", vertical="center")
                sheet.cell(row= num+3 , column = 3+hours).alignment = Alignment(horizontal="center", vertical="center")
            const = str(2+ceil(len(group_list)/2)+3)
            gr1 = 'A'+ const
            sheet[gr1].value = "№"
            sheet[gr1].alignment = Alignment(horizontal="center", vertical="center")
            gr1 = 'B'+ const
            sheet[gr1].value = "ФИО"
            sheet[gr1].alignment = Alignment(horizontal="center", vertical="center")
            dim = 'A'+ str(int(const)-1) +':'+chr(67+hours) + str(int(const)-1)
            sheet.merge_cells(dim) 
            sheet.cell(row = int(const)-1, column = 1).value = self.group[j]+" "+self.list[j]+'ЛР'
            sheet['A'+ str(int(const)-1)].alignment = Alignment(horizontal="center", vertical="center")
            M = chr(67+hours) + str(int(const))
            sheet[M].value = "M%"
            sheet[M].alignment = Alignment(horizontal="center", vertical="center")  
            sheet.column_dimensions[chr(67+hours)].width = int(5)
            Begindate = datetime.strptime(self.data[j][0], "%d-%m-%Y" ) + timedelta(days=7)
            Enddate = Begindate
            for i in range(hours):
                M = chr(67+i)+ const
                sheet[M].value = str(Enddate.strftime("%d/%m"))[0:10]
                sheet[M].alignment = Alignment(horizontal="center", vertical="center", textRotation=90)
                sheet.column_dimensions[chr(67+i)].width = int(5)
                # sheet.row_dimensions[const].height  = int(42)
                
                Enddate = Begindate + timedelta(days=7)
                Begindate = Enddate 
            const = str(num-1) 
            for num in range(ceil(len(group_list)/2),len(group_list)):
                sheet.cell(row= num+1+int(const) , column = 1).value = num+1
                sheet.cell(row= num+1+int(const)  , column = 2).value = group_list[num]
                # summ = '= SUM(C3:' + chr(67+hours) + '3)'
                summ = '= 100*COUNTA(C'+str(num+1+int(const)) +':' + chr(67+hours-1) + str(num+1+int(const))+')'+'/'+str(hours)
                sheet.cell(row= num+1+int(const)  , column = 3+hours).value = summ
                sheet.cell(row= num+1+int(const)  , column = 1).alignment = Alignment(horizontal="center", vertical="center")
                sheet.cell(row= num+1+int(const)  , column = 2).alignment = Alignment(horizontal="center", vertical="center")
                sheet.cell(row= num+1+int(const)  , column = 3+hours).alignment = Alignment(horizontal="center", vertical="center")
        # wb.save("sample.xlsx")

    def creat_list_sr(self, wb, j):
        filename = self.group[j] +'.txt'
        group_list, group_dir = Info.creat_groups(filename)
        sheet = wb.create_sheet(str(self.group[j]) +'(сем+лр)')
        Info.beatifule(sheet)
        if self.hours[j][1] == 51:  hours_sem = 17
        elif self.hours[j][1]  == 17:   
            hours_sem = 8
        elif self.hours[j][1]  == 34:   hours_sem = 8
        if self.hours[j][2] == 51:  hours_lr = 17
        elif self.hours[j][2]  == 17:   hours_lr = 8
        elif self.hours[j][2]  == 34:   hours_lr = 8
        dim = 'A1:'+chr(67+hours_sem)+'1'
        sheet.merge_cells(dim) 
        sheet.cell(row = 1, column = 1).value = self.group[j]+" "+self.list[j]+'сем'
        sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
        M = chr(67+hours_sem)+'2'
        sheet[M].value = "M%"
        sheet[M].alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[chr(67+hours_sem)].width = int(5)
        Begindate = datetime.strptime(self.data[j][0], "%d-%m-%Y" )
        Enddate = Begindate
        for i in range(hours_sem):
            M = chr(67+i)+'2'
            sheet[M].value = str(Enddate.strftime("%d/%m"))[0:10]
            sheet[M].alignment = Alignment(horizontal="center", vertical="center", textRotation=90)
            sheet.column_dimensions[chr(67+i)].width = int(5)
            # sheet.row_dimensions['2'].height  = int(42)
            Enddate = Begindate + timedelta(days=7)
            Begindate = Enddate
        for num in range(len(group_list)): 
            sheet.cell(row= num+3 , column = 1).value = num+1
            sheet.cell(row= num+3 , column = 2).value = group_list[num]
            # summ = '= SUM(C3:' + chr(67+hours_sem-1) + '3)'
            summ = '= 100*COUNTA(C'+str(num+3) +':' + chr(67+hours_sem-1) + str(num+3)+')'+'/'+str(hours_sem)
            sheet.cell(row= num+3 , column = 3+hours_sem).value = summ
            sheet.cell(row= num+3 , column = 1).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row= num+3 , column = 2).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row= num+3 , column = 3+hours_sem).alignment = Alignment(horizontal="center", vertical="center")

        const = str(len(group_list)+5)
        gr1 = 'A'+ const
        sheet[gr1].value = "№"
        sheet[gr1].alignment = Alignment(horizontal="center", vertical="center")
        gr1 = 'B'+ const
        sheet[gr1].value = "ФИО"
        sheet[gr1].alignment = Alignment(horizontal="center", vertical="center")
        dim = 'A'+ str(int(const)-1) +':'+chr(67+hours_lr) + str(int(const)-1)
        sheet.merge_cells(dim) 
        sheet.cell(row = int(const)-1, column = 1).value = self.group[j]+" "+self.list[j]+'ЛР'
        sheet['A'+ str(int(const)-1)].alignment = Alignment(horizontal="center", vertical="center")
        M = chr(67+hours_lr) + str(int(const))
        sheet[M].value = "M%"
        sheet[M].alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[chr(67+hours_lr)].width = int(5)
        Begindate = datetime.strptime(self.data[j][1], "%d-%m-%Y" )
        Enddate = Begindate
        for i in range(hours_lr):
            M = chr(67+i)+ const
            sheet[M].value = str(Enddate.strftime("%d/%m"))[0:10]
            sheet[M].alignment = Alignment(horizontal="center", vertical="center", textRotation=90)
            sheet.column_dimensions[chr(67+i)].width = int(5)
            # sheet.row_dimensions[const].height  = int(42)
            Enddate = Begindate + timedelta(days=7)
            Begindate = Enddate
        for num in range(ceil(len(group_list)/2)): 
            sheet.cell(row= num+1+int(const) , column = 1).value = num+1
            sheet.cell(row= num+1+int(const)  , column = 2).value = group_list[num]
            # summ = '= SUM(C3:' + chr(67+hours_lr-1) + '3)'
            summ = '= 100*COUNTA(C'+str(num+1+int(const) ) +':' + chr(67+hours_lr-1) + str(num+1+int(const) )+')'+'/'+str(hours_lr)
            sheet.cell(row= num+1+int(const)  , column = 3+hours_lr).value = summ
            sheet.cell(row= num+1+int(const)  , column = 1).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row= num+1+int(const)  , column = 2).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row= num+1+int(const)  , column = 3+hours_lr).alignment = Alignment(horizontal="center", vertical="center")
        const = str(len(group_list)+5+ceil(len(group_list)/2)+3)
        gr1 = 'A'+ const
        sheet[gr1].value = "№"
        sheet[gr1].alignment = Alignment(horizontal="center", vertical="center")
        gr1 = 'B'+ const
        sheet[gr1].value = "ФИО"
        sheet[gr1].alignment = Alignment(horizontal="center", vertical="center")
        dim = 'A'+ str(int(const)-1) +':'+chr(67+hours_lr) + str(int(const)-1)
        sheet.merge_cells(dim) 
        sheet.cell(row = int(const)-1, column = 1).value = self.group[j]+" "+self.list[j]+'ЛР'
        sheet['A'+ str(int(const)-1)].alignment = Alignment(horizontal="center", vertical="center")
        M = chr(67+hours_lr) + str(int(const))
        sheet[M].value = "M%"
        sheet[M].alignment = Alignment(horizontal="center", vertical="center")   
        sheet.column_dimensions[chr(67+hours_lr)].width = int(5)
        Begindate = datetime.strptime(self.data[j][1], "%d-%m-%Y" ) + timedelta(days=7)
        Enddate = Begindate
        for i in range(hours_sem):
            M = chr(67+i)+ const
            sheet[M].value = str(Enddate.strftime("%d/%m"))[0:10]
            sheet[M].alignment = Alignment(horizontal="center", vertical="center", textRotation=90)
            sheet.column_dimensions[chr(67+i)].width = int(5)
            # sheet.row_dimensions[const].height  = int(42)
            Enddate = Begindate + timedelta(days=7)
            Begindate = Enddate
        const = str(len(group_list)+5+3) 
        for num in range(ceil(len(group_list)/2),len(group_list)):
            sheet.cell(row= num+1+int(const) , column = 1).value = num+1
            sheet.cell(row= num+1+int(const)  , column = 2).value = group_list[num]
            # summ = '= SUM(C3:' + chr(67+hours_lr-1) + '3)'
            summ = '= 100*COUNTA(C'+str(num+1+int(const) ) +':' + chr(67+hours_lr-1) + str(num+1+int(const) )+')'+'/'+str(hours_lr)
            sheet.cell(row= num+1+int(const)  , column = 3+hours_lr).value = summ
            sheet.cell(row= num+1+int(const)  , column = 1).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row= num+1+int(const)  , column = 2).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row= num+1+int(const)  , column = 3+hours_lr).alignment = Alignment(horizontal="center", vertical="center")
        # wb.save("sample.xlsx")

    def creat_groups(filename):
            Group_list =[]
            with open(filename, 'r', encoding='UTF-8') as file:
                while line := file.readline():
                    tmp=line.rstrip().split()
                    if len(tmp)==3:
                        tmp = tmp[0]+'_'+tmp[1]+tmp[2]
                    else:
                        tmp = tmp[0]+'_'+tmp[1]
                    Group_list.append(tmp)
                    
            group_dir = filename[:-4]  
            return Group_list, group_dir

class Do(Info):
    def __init__(self):
        global listt
        global group
        global hours 
        global data
        self.list = listt
        self.group = group
        self.hours = hours
        self.data = data
        wb = "sample.xlsx"
        Do.extract_text(self, wb)

    def extract_text(self, wb):
        cwd = os.getcwd()
        base_path = os.path.join(cwd,'attachment')
        dir_list = os.listdir(base_path) 
        
        for dir in dir_list:  
            source = os.path.join(base_path, dir)
            get_files = os.listdir(source)
            for file in get_files:
                pdf_path = os.path.join(source, file)
                first_page = Do.extract_text_by_page(pdf_path)
                Do.determine_where(self, file, dir, first_page, wb)
                
            old_path = os.path.join(cwd,'attachment', dir)
            shutil.rmtree(old_path)

    def extract_text_by_page(pdf_path):
            with open(pdf_path, 'rb') as fh:
                for page in PDFPage.get_pages(fh, caching=True, check_extractable=True):
                    resource_manager = PDFResourceManager()
                    fake_file_handle = io.StringIO()
                    converter = TextConverter(resource_manager, fake_file_handle)
                    page_interpreter = PDFPageInterpreter(resource_manager, converter)
                    page_interpreter.process_page(page)
                    text = fake_file_handle.getvalue()
                    # yield text
                    converter.close()
                    fake_file_handle.close()
                    return text

    def determine_where(self, filename, dir, first_page, wb):
        f = filename
        cwd = os.getcwd()
        for i in range(len(self.list)):
            if self.list[i] in first_page:
                if self.group[i] in first_page:
                    old_path = os.path.join(cwd,'attachment', dir)
                    old_file = os.path.join(cwd,'attachment', dir, f)
                    new_path = os.path.join(cwd, self.list[i][:-1], self.group[i])
                    new_path_full = os.path.join(cwd, self.list[i][:-1], self.group[i],  dir)
                    new_file = os.path.join(cwd, self.list[i][:-1], self.group[i],  dir, f)
                    
                    if os.path.exists(new_path):
                        pass 
                    else:
                        os.mkdir(new_path)

                    if os.path.exists(new_path_full):
                        pass 
                    else:
                        os.mkdir(new_path_full)

                    if os.path.exists(new_file):
                        os.remove(new_file)   
                    
                    shutil.move(old_file, new_file)
                    Do.mark_table(self, wb, first_page, i, dir)
                    # shutil.rmtree(old_file)

    def find_column_row(self,wb,search_text):
        resulution = None
        wb = openpyxl.load_workbook(wb)
        sheets_list = wb.sheetnames
        for i in range(len(sheets_list)):
            # if search_text == 'ЛР':
            #     sheet_active = wb[sheets_list[page]]
            # else: sheet_active = wb[sheets_list[i]]
            sheet_active = wb[sheets_list[i]]
            row_max = sheet_active.max_row  # Получаем количество столбцов
            column_max = sheet_active.max_column 
            row_min = 1 #Переменная, отвечающая за номер строки
            column_min = 1 #Переменная, отвечающая за номер столбца
            while column_min <= column_max:
                row_min_min = row_min
                row_max_max = row_max
                while row_min_min <= row_max_max:
                    row_min_min = str(row_min_min)
                    word_column = get_column_letter(column_min)
                    word_column = str(word_column)
                    word_cell = word_column + row_min_min
                    data_from_cell = sheet_active[word_cell].value
                    data_from_cell = str(data_from_cell)
                    regular = search_text
                    result = re.findall(regular, data_from_cell)
                    if len(result) > 0:
                        # print('Нашли в ячейке:', word_cell)
                        resulution = word_cell
                        page_return = i
                    row_min_min = int(row_min_min)
                    row_min_min = row_min_min + 1
                column_min = column_min + 1
        # if search_text == 'ЛР':
        #     return resulution
        # else: return resulution, page_return
        return resulution, page_return
        
    def mark_table(self, wb, first_page, i, dir):
        if 'ЛАБ' in first_page:
            search_text = dir
            # page = 1
            word_cell, page = Do.find_column_row(self,wb,search_text)
            # print(dir, word_cell, page)
            # search_text = 'ЛР'
            # lr_cell = Do.find_column_row(self,wb,search_text, page)
            # print(lr_cell)
            num = first_page[first_page.find('№')+1]
            wb = openpyxl.load_workbook(wb)
            sheets_list = wb.sheetnames
            sheet_active = wb[sheets_list[page]]
            M = chr(66+int(num))+word_cell[1:]
            print(M)
            sheet_active[M].alignment = Alignment(horizontal="center", vertical="center")
            sheet_active[M].value = int('1')
            wb.save("sample.xlsx")

if __name__ == '__main__':
    filename = 'Дисциплины.txt'
    global listt
    global group
    global hours 
    global data
    inf = Info()
    do = Do()
    print()
